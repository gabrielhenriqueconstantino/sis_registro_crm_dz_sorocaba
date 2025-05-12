import tkinter as tk
from tkinter import font
import customtkinter as ctk
from tkinter import font as tkfont
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import locale
import sqlite3
from datetime import datetime
import time
import requests  # Para integração com a API de CEP
import re
from PIL import Image, ImageTk
from PIL import ImageOps
import os
import sys
import subprocess
import platform
import re
import threading

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
import os

from screeninfo import get_monitors
from analise_dados import JanelaAnalise
from pathlib import Path
import webbrowser

from analise_dados import JanelaAnalise

class SistemaProtocolos:
    def __init__(self, root):
        self.root = root
        self.root.attributes('-alpha', 0.98)
        self.root.title("Sistema de Registro, Consulta e Análise de Protocolos de Reclamação")

        # Caminho para o ícone
        caminho_icone = Path("sis_crm/img/icons/icons_window/logo_sis_crm.ico")
        
        # Definir o ícone - chamando iconbitmap no root (objeto Tk)
        try:
            self.root.iconbitmap(caminho_icone)
        except Exception as e:
            print(f"Icone não encontrado: {e}")

        # Define tamanho da janela
        largura_janela = 1366
        altura_janela = 768

        # Calcula a posição central
        largura_tela = self.root.winfo_screenwidth()
        altura_tela = self.root.winfo_screenheight()

        x = (largura_tela // 2) - (largura_janela // 2)
        y = (altura_tela // 2) - (altura_janela // 2)

        # Aplica o tamanho + posição centralizada
        self.root.geometry(f"{largura_janela}x{altura_janela}+{x}+{y}")

        self.janela_analise = None  # Inicialmente não tem nenhuma janela aberta
        
        self._em_animacao = False

        #Cria o frame principal
        self.frame_adicionar_crm = ctk.CTkFrame(root,  border_color="white", border_width=1)
        self.frame_adicionar_crm.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

        self.frame_exibir_pedidos = ctk.CTkFrame(root, border_color="white", border_width=1)
        self.frame_exibir_pedidos.grid(row=2, column=0, padx=10, pady=5, sticky="nsew")

        # Configurar o layout da janela principal
        root.grid_rowconfigure(0, weight=1)
        root.grid_rowconfigure(1, weight=1)
        root.grid_rowconfigure(2, weight=5)
        root.grid_columnconfigure(0, weight=1)

        # Adicione no início do seu código (após criar a janela principal)
        style = ttk.Style()
        style.configure('TCombobox', fieldbackground='white', background='white')
        style.configure('TEntry', fieldbackground='white')

        self.criar_widgets()
        
        # Exibir todos os protcolos ao iniciar o aplicativo
        self.exibir_protocolos()

    def criar_widgets(self):
        self.criar_frame_logo()
        self.criar_frame_adicionar_crm()
        self.criar_frame_exibir_pedidos()

    @staticmethod
    def obter_resolucao():
        monitor = get_monitors()[0]
        return monitor.width, monitor.height

    def resolucao(self):
        largura, altura = self.obter_resolucao()
        print(f'Sua resolução é {largura}x{altura}')

    def carregar_logo_dinamica(self):
        tema = ctk.get_appearance_mode()  # Retorna "Light" ou "Dark"
    
        caminho_base = Path("sis_crm/img/logo")
    
        if tema == "Light":
            imagem_logo = Image.open(caminho_base / "logo_preto.png")
        else:
            imagem_logo = Image.open(caminho_base / "logo_branco.png")

        # Obtém a resolução do monitor
        largura, altura = SistemaProtocolos.obter_resolucao()

        # Define o tamanho padrão
        largura_logo = 350
        altura_logo = 90

        # Se for monitor 1024x768, muda o tamanho da logo
        if largura == 1024 and altura == 768:
            largura_logo = 200  # menor largura
            altura_logo = int(largura_logo * (90/350))  # mantém a mesma proporção original

        # Redimensiona a imagem
        imagem_logo = imagem_logo.resize((largura_logo, altura_logo), Image.Resampling.LANCZOS)

        # Usa CTkImage para suporte a HighDPI
        self.logo_img = ctk.CTkImage(dark_image=imagem_logo, light_image=imagem_logo, size=(largura_logo, altura_logo))
        self.logo_label.configure(image=self.logo_img)


    def criar_frame_logo(self):
        # Pega a resolução do monitor
        largura, altura = SistemaProtocolos.obter_resolucao()

        self.frame_logo = ctk.CTkFrame(self.root, bg_color="transparent", border_color="white", border_width=1)
        self.frame_logo.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        self.frame_logo.grid_rowconfigure(0, weight=1)  # Espaço acima
        self.frame_logo.grid_rowconfigure(1, weight=0)  # Linha dos widgets
        self.frame_logo.grid_rowconfigure(2, weight=1)  # Espaço abaixo

        self.frame_logo.grid_columnconfigure(0, weight=1)  # Logo
        self.frame_logo.grid_columnconfigure(1, weight=0)  # Espaço entre logo e opções
        self.frame_logo.grid_columnconfigure(2, weight=0)  # frame_opcoes
        self.frame_logo.grid_columnconfigure(3, weight=0)  # Espaço entre opções e opções2
        self.frame_logo.grid_columnconfigure(4, weight=0)  # frame_opcoes2
        self.frame_logo.grid_columnconfigure(5, weight=1)  # Espaço final para centralizar tudo

        # Define o PADX dinamicamente:
        if largura == 1024 and altura == 768:
            padding_logo = (10, 15)  # Muito menos espaço lateral
        else:
            padding_logo = (10, 150)  # Seu padrão atual

        # Label da logo (a imagem será definida depois)
        self.logo_label = ctk.CTkLabel(self.frame_logo, text="")
        self.logo_label.grid(row=1, column=0, padx=padding_logo, pady=5, sticky="w")

        # Carrega a imagem correta com base no tema
        self.carregar_logo_dinamica()

        # Frame de botões principais (editar, excluir, gráfico)
        self.frame_opcoes = ctk.CTkFrame(self.frame_logo, border_color="white", border_width=1)
        self.frame_opcoes.grid(row=1, column=2, padx=10, pady=10, sticky="nsew", ipadx=80)

        # Frame de botões secundários (tema, reload, info)
        self.frame_opcoes2 = ctk.CTkFrame(self.frame_logo, border_color="white", border_width=1)
        self.frame_opcoes2.grid(row=1, column=4, padx=(50, 10), pady=10, sticky="nsew")

        # Botões dentro de frame_opcoes
        for col in range(3):
            self.frame_opcoes.grid_columnconfigure(col, weight=1)

        self.icone_editar = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/pen.png"), size=(30, 30))
        ctk.CTkButton(self.frame_opcoes,
                      image=self.icone_editar, 
                      text="", 
                      font=("Arial", 29), 
                      width=60, 
                      height=60,
                      fg_color="#0078D7",
                      hover_color="#005A9E",  # cor mais escura para o hover
                      text_color="white",
                      command=self.editar_protocolo, 
                      ).grid(row=0, 
                             column=0, 
                             padx=10, 
                             pady=20)
        
        self.icone_excluir = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/delete.png"), size=(30, 30))
        ctk.CTkButton(self.frame_opcoes,
                      image=self.icone_excluir, 
                      text="", 
                      font=("Arial", 30), 
                      width=60, 
                      height=60,
                      fg_color="#0078D7",
                      hover_color="#005A9E",  # cor mais escura para o hover
                      text_color="white",
                      command=self.excluir_protocolo
                      ).grid(row=0, 
                             column=1, 
                             padx=10, 
                             pady=20)

        self.icone_dados = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/data.png"), size=(30, 30))              
        ctk.CTkButton(self.frame_opcoes,
                      image=self.icone_dados, 
                      text="", 
                      font=("Arial", 30), 
                      width=60, 
                      height=60,
                      fg_color="#0078D7",
                      hover_color="#005A9E",  # cor mais escura para o hover
                      text_color="white",
                      command=self.abrir_janela_analise, 
                      ).grid(row=0, 
                      column=2, 
                      padx=10, 
                      pady=20)

        # Botões dentro de frame_opcoes2
        for col in range(3):
            self.frame_opcoes2.grid_columnconfigure(col, weight=1)

        self.icone_lua = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/moon.png"), size=(35, 35))
        self.icone_sol = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/sun.png"), size=(30, 30))
        self.icone_inicial = self.icone_sol if ctk.get_appearance_mode() == "Dark" else self.icone_lua
        self.botao_tema = ctk.CTkButton(self.frame_opcoes2, 
                                        text="",
                                        image=self.icone_inicial, 
                                        font=("Arial", 30),  
                                        width=60, 
                                        height=60,
                                        fg_color="#0078D7",
                                        hover_color="#005A9E",  # cor mais escura para o hover
                                        text_color="white",
                                        command=self.alternar_tema)
        self.botao_tema.grid(row=0, column=0, padx=20, pady=20)

        self.icone_reload = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/return.png"), size=(20, 20))
        self.botao_reload = ctk.CTkButton(self.frame_opcoes2, 
                                          image=self.icone_reload,
                                          text="", 
                                          font=("Arial", 30), 
                                          width=60, 
                                          height=60,
                                          fg_color="#0078D7",
                                          hover_color="#005A9E",  # cor mais escura para o hover
                                          text_color="white",
                                          command=self.confirmar_reinicio)
        self.botao_reload.grid(row=0, column=1, padx=20, pady=20)

        self.icone_info = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/info.png"), size=(25, 25))
        self.botao_info = ctk.CTkButton(self.frame_opcoes2,
                                        image=self.icone_info, 
                                        text="", 
                                        font=("Arial", 30), 
                                        width=60, 
                                        height=60,
                                        fg_color="#0078D7",
                                        hover_color="#005A9E",  # cor mais escura para o hover
                                        text_color="white",
                                        command=self.abrir_janela_info)
        self.botao_info.grid(row=0, column=2, padx=20, pady=20)


    def alternar_tema(self):
        """Versão com transição suave"""
        if not self._em_animacao:  # Previne múltiplas animações simultâneas
            self._em_animacao = True
        self._iniciar_animacao_tema()

    def _iniciar_animacao_tema(self):
        """Controla a animação completa"""
        # Fade Out
        self._animar_alpha(0.98, 0.2, 15, lambda: self._trocar_tema_efetivo())
        
    def _animar_alpha(self, inicio, fim, passos, callback=None):
        """Animação genérica de alpha"""
        step = (fim - inicio) / passos
        current = inicio
        
        def _step():
            nonlocal current
            current += step
            if (step > 0 and current >= fim) or (step < 0 and current <= fim):
                current = fim
                self.root.attributes('-alpha', current)
                if callback:
                    callback()
                return
            
            self.root.attributes('-alpha', current)
            self.root.after(10, _step)
        
        _step()

    def _trocar_tema_efetivo(self):
        """Troca o tema e inicia fade in"""
        tema_atual = ctk.get_appearance_mode()
        novo_tema = "Dark" if tema_atual == "Light" else "Light"
        ctk.set_appearance_mode(novo_tema)
        
        # Atualiza elementos visuais
        self.carregar_logo_dinamica()
        self.aplicar_tema_treeview()
        novo_icone = self.icone_sol if novo_tema == "Dark" else self.icone_lua
        self.botao_tema.configure(image=novo_icone)
        
        # Fade In
        self._animar_alpha(0.2, 0.98, 15, lambda: setattr(self, '_em_animacao', False))  
    
    #Aplicar tema claro ou escuro na treeview(tabela)
    def aplicar_tema_treeview(self):
        estilo = ttk.Style()
        tema = ctk.get_appearance_mode()

        estilo.theme_use("clam")  # Tema neutro personalizável

        if tema == "Dark":
            estilo.configure("Treeview",
                         background="#2b2b2b",
                         foreground="white",
                         fieldbackground="#2b2b2b",
                         rowheight=31,
                         bordercolor="#444",
                         borderwidth=1,
                         font=("Poppins", 10))
            estilo.map("Treeview", background=[('selected', '#444444')])

            estilo.configure("Treeview.Heading",
                         background="#1e1e1e",
                         foreground="white",
                         relief="flat",
                         font=("Poppins", 10, "bold"),
                         borderwidth=2,
                         bordercolor="white")  # Borda branca ao redor dos headings
            estilo.map("Treeview.Heading", background=[("!active", "#1e1e1e")])

        else:
            estilo.configure("Treeview",
                         background="white",
                         foreground="black",
                         fieldbackground="white",
                         rowheight=31,
                         bordercolor="#ccc",
                         borderwidth=2,
                         font=("Poppins", 10))
            estilo.map("Treeview", background=[('selected', '#0078D7')], foreground=[('selected', 'white')])

            estilo.configure("Treeview.Heading",
                         background="#e1e1e1",
                         foreground="black",
                         relief="flat",
                         font=("Poppins", 10, "bold"),
                         borderwidth=2,
                         bordercolor="black")  # Borda preta para modo claro
            estilo.map("Treeview.Heading", background=[("!active", "#e1e1e1")])

        self.exibir_protocolos()

    #Valida o boletim para não incluir letras, simbolos e inteiros maior do que 6 digitos
    def validar_boletim_evento(self, event):
        valor = self.entry_boletim.get()
        if valor and (not valor.isdigit() or len(valor) > 6):
            # Remove o último caractere inválido
            self.entry_boletim.delete(len(valor)-1, tk.END)

    #Validar protocolos para não incluir mais de 16 caracteres, contando espaços tbm.
    def validar_protocolo_evento(self, event):
        valor = self.entry_protocolo.get()
    
        # Remove espaços e converte para maiúsculas
        novo_valor = valor.replace(" ", "").upper()
    
        # Verifica se o valor foi alterado (havia espaços)
        if novo_valor != valor:
            self.entry_protocolo.delete(0, tk.END)
            self.entry_protocolo.insert(0, novo_valor)
            return
    
        # Verifica se excedeu o tamanho máximo
        if len(novo_valor) > 16:
            self.entry_protocolo.delete(16, tk.END)
            return
    
        # Verifica caracteres inválidos (permite letras, números e hífen)
        for char in novo_valor:
            if not (char.isalnum() or char == '-'):
                # Remove o último caractere inválido
                self.entry_protocolo.delete(len(novo_valor)-1, tk.END)
                return

    def criar_frame_adicionar_crm(self):
        ## Configurar colunas para centralização
        for i in range(7):
            self.frame_adicionar_crm.grid_columnconfigure(i, weight=1)

            self.frame_adicionar_crm.grid_rowconfigure(0, weight=1)
            self.frame_adicionar_crm.grid_rowconfigure(1, weight=1)

            # Estilo para harmonizar calendario com o CustomTkinter
            style = ttk.Style()
            style.theme_use('clam')  # Tema neutro que permite customizações
            style.configure("Custom.DateEntry",
                    fieldbackground="#2b2b2b",  # Fundo interno
                    background="#2b2b2b",      # Fundo do botão de calendário
                    foreground="white",        # Cor do texto
                    arrowcolor="white",        # Cor da setinha
                    bordercolor="#3a3a3a",      # Cor da borda
                    lightcolor="#3a3a3a",
                    darkcolor="#3a3a3a")

            # Campo de data com visual mais escuro
            self.data_entry = DateEntry(self.frame_adicionar_crm,
                            date_pattern='dd/mm/yyyy',
                            style="Custom.DateEntry",
                            borderwidth=1,
                            font=("Arial", 10),
                            justify="center",
                            locale='pt_BR')
            self.data_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

            # Campo de data com calendário (DateEntry)
            self.data_entry = DateEntry(self.frame_adicionar_crm, date_pattern='dd/mm/yyyy', locale="pt_BR")
            self.data_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
            
            # Entrada para boletim (somente números inteiros, até 6 dígitos)
            self.entry_boletim = ctk.CTkEntry(
            self.frame_adicionar_crm,
            placeholder_text="Boletim"
            )
            self.entry_boletim.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
            self.entry_boletim.bind("<KeyRelease>", self.validar_boletim_evento)

            # Entrada para protocolo
            self.entry_protocolo = ctk.CTkEntry(self.frame_adicionar_crm, placeholder_text="Protocolo")
            self.entry_protocolo.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
            self.entry_protocolo.bind("<KeyRelease>", self.validar_protocolo_evento)

            # ComboBox para a situação
            self.combo_situacao = ctk.CTkComboBox(self.frame_adicionar_crm, values=[
                "Situação (selecione)",
                "Aberto",
                "Reaberto",
                "Em andamento",
                "Em retorno",
                "Resolvido"
                ], state="readonly")
            self.combo_situacao.set("Situação (selecione)")
            self.combo_situacao.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

            # ComboBox para a área municipal
            self.combo_area = ctk.CTkComboBox(self.frame_adicionar_crm, values=[
            "Área (selecione)", 
            "Sudoeste", 
            "Centro-Sul", 
            "Centro-Norte", 
            "Noroeste", 
            "Norte", 
            "Leste"], state="readonly")
            self.combo_area.set("Área (selecione)")
            self.combo_area.grid(row=0, column=4, padx=5, pady=5, sticky="ew")

            # ComboBox para os problemas
            self.combo_problema = ctk.CTkComboBox(self.frame_adicionar_crm, values=[
                "Problema (selecione)",
                "Focos de dengue",
                "Imóveis abandonados",
                "Acúmulo de lixo e materiais inservíveis",
                "Mato alto/entulho",
                "Animais Peçonhentos",
                "Pombos",
                "Morcegos",
                "Infestação de Roedores",
                "Carrapatos, pulgas e outros ectoparasitas",
                "Caramujos",
                "Criação de animais rurais em zona urbana",
                "Outras Reclamações de Saúde Pública"
            ], state="readonly")
            self.combo_problema.set("Problema (selecione)")
            self.combo_problema.grid(row=0, column=5, padx=5, pady=5, sticky="ew")

            # Entrada para observações
            self.entry_observacoes = ctk.CTkEntry(self.frame_adicionar_crm, placeholder_text="Observações(se houver)")
            self.entry_observacoes.grid(row=0, column=6, padx=5, pady=5, sticky="ew")

        # Botão para adicionar na tabela
        ctk.CTkButton(
        self.frame_adicionar_crm,
        text="SALVAR",
        command=self.adicionar_protocolo,
        width=350,
        fg_color="#0078D7",
        hover_color="#005A9E",  # azul mais escuro no hover
        text_color="white",
        font=("Poppins", 14, "bold")  # você pode ajustar o tamanho ou peso
        ).grid(
        row=1,
        column=0,
        columnspan=7,
        pady=10,
        padx=10
        )

    #BANCO DE DADOS    
    def criar_banco_dados():
        caminho_banco = Path("sis_crm/database/db/sistema_protocolos.db")
        conn = sqlite3.connect(caminho_banco)  # Alterado o nome do banco
        cursor = conn.cursor()

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS protocolos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_abertura TEXT NOT NULL,
            boletim INTERGER NOT NULL,
            protocolo TEXT NOT NULL,
            situacao TEXT NOT NULL,
            area TEXT NOT NULL,
            bairro TEXT,
            problema TEXT,
            observacao TEXT,
            aberto_por TEXT
            )
        ''')

        conn.commit()
        conn.close()
    
    # Chamar a função ao iniciar a aplicação
    criar_banco_dados()

    def criar_frame_exibir_pedidos(self):
        largura, altura = SistemaProtocolos.obter_resolucao()

        ctk.CTkLabel(
        self.frame_exibir_pedidos,
        text="Protocolos de reclamação",
        font=("Poppins", 14, "bold"),
        text_color=("black", "white")
        ).grid(row=0, column=0, columnspan=2, pady=(0, 15))

        self.entry_pesquisa_principal = ctk.CTkEntry(self.frame_exibir_pedidos, 
                                           placeholder_text="Pesquise por data, boletim, área, situação, problema, aberto por...", 
                                           corner_radius=12, width=750, 
                                           border_color="white", border_width=1)
        self.entry_pesquisa_principal.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        self.entry_pesquisa_principal.bind("<KeyRelease>", self.atualizar_pesquisa_geral)

        self.entry_pesquisa_id = ctk.CTkEntry(self.frame_exibir_pedidos, 
                                    placeholder_text="Pesquise exclusivamente por ID ou PROTOCOLO", 
                                    corner_radius=12, width=400, 
                                    border_color="white", border_width=1)
        self.entry_pesquisa_id.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        self.entry_pesquisa_id.bind("<KeyRelease>", self.atualizar_pesquisa_protocolo)

        # Configuração do grid para expansão
        self.frame_exibir_pedidos.grid_rowconfigure(2, weight=1)
        self.frame_exibir_pedidos.grid_columnconfigure(0, weight=1)
        self.frame_exibir_pedidos.grid_columnconfigure(1, weight=1)

        frame_treeview = ctk.CTkFrame(self.frame_exibir_pedidos)
        frame_treeview.grid(row=2, column=0, columnspan=2, padx=25, pady=10, sticky="nsew")
    
        # Configuração do grid dentro do frame_treeview
        frame_treeview.grid_rowconfigure(0, weight=1)
        frame_treeview.grid_columnconfigure(0, weight=1)
        frame_treeview.grid_columnconfigure(1, weight=0)  # Para a scrollbar

        # Criação da Treeview
        self.treeview_protocolos = ttk.Treeview(
            frame_treeview,
            columns=("ID", "Data Abertura", "Boletim", "Protocolo", "Situação", 
                "Área", "Bairro", "Problema", "Observação", "Aberto por"),
            show='headings'
        )

        # Configuração das colunas 
        self.treeview_protocolos.heading("ID", text="ID")
        self.treeview_protocolos.heading("Data Abertura", text="Data")
        self.treeview_protocolos.heading("Boletim", text="Boletim")
        self.treeview_protocolos.heading("Protocolo", text="Protocolo")
        self.treeview_protocolos.heading("Situação", text="Situação")
        self.treeview_protocolos.heading("Área", text="Área")
        self.treeview_protocolos.heading("Bairro", text="Bairro")
        self.treeview_protocolos.heading("Problema", text="Problema")
        self.treeview_protocolos.heading("Observação", text="Observação")
        self.treeview_protocolos.heading("Aberto por", text="Aberto por")

        # Ajuste das larguras das colunas para 1024x768
        if largura == 1024 and altura == 768:
            self.treeview_protocolos.column("ID", width=20)
            self.treeview_protocolos.column("Data Abertura", width=80, anchor="center")
            self.treeview_protocolos.column("Boletim", width=70, anchor="center")
            self.treeview_protocolos.column("Protocolo", width=100, anchor="center")
            self.treeview_protocolos.column("Situação", width=100, anchor="center")
            self.treeview_protocolos.column("Área", width=80, anchor="center")
            self.treeview_protocolos.column("Bairro", width=100)
            self.treeview_protocolos.column("Problema", width=120)
            self.treeview_protocolos.column("Observação", width=80)
            self.treeview_protocolos.column("Aberto por", width=80, anchor="center")
        else:
            self.treeview_protocolos.column("ID", width=30)
            self.treeview_protocolos.column("Data Abertura", width=100, anchor="center")
            self.treeview_protocolos.column("Boletim", width=80, anchor="center")
            self.treeview_protocolos.column("Protocolo", width=125, anchor="center")
            self.treeview_protocolos.column("Situação", width=145, anchor="center")
            self.treeview_protocolos.column("Área", width=120, anchor="center")
            self.treeview_protocolos.column("Bairro", width=170)
            self.treeview_protocolos.column("Problema", width=160)
            self.treeview_protocolos.column("Observação", width=120)
            self.treeview_protocolos.column("Aberto por", width=120, anchor="center")

        # Scrollbar e Treeview usando grid (melhor para controle)
        self.scrollbar_vertical = ctk.CTkScrollbar(frame_treeview, orientation="vertical", command=self.treeview_protocolos.yview)
        self.treeview_protocolos.configure(yscrollcommand=self.scrollbar_vertical.set)

        # Usar grid em vez de pack para melhor controle
        self.treeview_protocolos.grid(row=0, column=0, sticky="nsew")
        self.scrollbar_vertical.grid(row=0, column=1, sticky="ns")

        self.treeview_protocolos.bind("<Delete>", lambda event: self.excluir_protocolo())
        self.treeview_protocolos.bind("<Double-1>", self.editar_protocolo)

        self.exibir_protocolos()
        self.aplicar_tema_treeview()

    #busca o nome do usuário do sistema operacional
    def obter_usuario_logado(self):
        """Retorna o nome do usuário logado de forma multiplataforma"""
        try:
            sistema = platform.system()
            if sistema == "Windows":
                return os.getlogin()  # Funciona no Windows
            else:
                return os.environ.get('USER', 'Desconhecido')  # Funciona no Linux/macOS
        except Exception:
            return "Desconhecido"  # Caso haja algum erro
    
    #buscar o protocolo no site da prefeitura e pegar o endereço e bairro para colocar na tabela
    def buscar_bairro(self):
        # Caminho para o ChromeDriver
        chromedriver_path = Path("sis_crm/chromedriver/chromedriver.exe")  # Ajuste o caminho se necessário

        # Configurações do navegador Chrome
        options = Options()
        options.headless = False  # Deixe como False para ver o navegador em ação

        # Criando o serviço e inicializando o driver
        service = Service(chromedriver_path)
        driver = webdriver.Chrome(service=service, options=options)

        # Acessando o site diretamente
        buscar_protocolo = self.entry_protocolo.get()  # Exemplo de protocolo
        url = f"http://central156.sorocaba.sp.gov.br/atendimento/#/User/Request?protocolo={buscar_protocolo}"
        driver.get(url)

        try:
            # Verificar se a página carregou completamente
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script('return document.readyState') == 'complete'
            )
        
            # Espera adicional para garantir que o conteúdo dinâmico foi carregado
            time.sleep(2)  # Ajuste este tempo conforme necessário

            # Espera explícita para o elemento estar presente
            # Localizar a div que contém o cabeçalho "Endereço" e então pegar o label dentro dela
            endereco_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, 
                '//div[@class="list-group-item"][.//h3[contains(text(), "Endereço")]]'
                '//div[@class="row"][.//div[@class="small-12 columns"]]'
                '//label[@class="ng-binding"]'))
            )

            # Destacar visualmente o elemento encontrado (seleção amarela)
            driver.execute_script(
            "arguments[0].style.backgroundColor = 'yellow';"
            "arguments[0].style.border = '2px solid red';"
            "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
            endereco_element
            )

            # Manter o destaque por 3 segundos para visualização
            time.sleep(4)

            # Remover o destaque
            driver.execute_script(
            "arguments[0].style.backgroundColor = '';"
            "arguments[0].style.border = '';",
            endereco_element
            )
        
            # Exibir o conteúdo completo do endereço
            endereco_completo = endereco_element.text
            print("Endereço completo:", endereco_completo)

            # Usar expressão regular para pegar o bairro e remover hifens
            bairro_match = re.search(r'(\D+)\s*-\s*Sorocaba\s*/SP$', endereco_completo)
            if bairro_match:
                bairro = bairro_match.group(1).strip()

                # Remover todos os hifens do bairro
                bairro = re.sub(r'-', '', bairro).strip()
            
                #especifico para a Vila Barão
                if bairro.lower() in ["barão", "barao"]:
                    bairro = "Vila Barão"

                print("Bairro encontrado:", bairro)
                return bairro  # Retornar o bairro encontrado
            else:
                print("Bairro não encontrado.")
                return None  # Retornar None caso não encontre o bairro

        except Exception as e:
            print(f"Erro: {e}")
            return None  # Retornar None em caso de erro

        finally:
            driver.quit()  # Garantir que o driver será fechado corretamente

    #adicionar protocolo na tabela
    def adicionar_protocolo(self):
        data_abertura = self.data_entry.get()
        boletim = self.entry_boletim.get()
        protocolo = self.entry_protocolo.get()
        situacao = self.combo_situacao.get()
        area = self.combo_area.get()

        bairro = self.buscar_bairro()
        if not bairro:
            bairro = "ERRO - Adicione manualmente"

        problema = self.combo_problema.get()
        observacao = self.entry_observacoes.get()
        aberto_por = self.obter_usuario_logado()

        # Verificar se algum campo obrigatório está vazio ou com valor padrão inválido
        if not all([data_abertura, boletim, protocolo, situacao, area, problema]) or \
            situacao == "Situação (selecione)" or \
            area == "Área (selecione)" or \
            problema == "Problema (selecione)":
            messagebox.showwarning("Atenção", "Preencha todos os campos obrigatórios corretamente.")
            return

        caminho_banco = Path("sis_crm/database/db/sistema_protocolos.db")
        conn = sqlite3.connect(caminho_banco)  # Alterado o nome do banco
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO protocolos (data_abertura, boletim, protocolo, situacao, area, bairro, problema, observacao, aberto_por)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (data_abertura, boletim, protocolo, situacao, area, bairro, problema, observacao, aberto_por))

        protocolo_id = cursor.lastrowid
        
        conn.commit()
        conn.close()

        # Backup automático para Excel
        caminho_excel = Path("sis_crm/database/backup/backup.xlsx")
        caminho_excel.parent.mkdir(parents=True, exist_ok=True)

        cabeçalhos = ["ID", "Data Abertura", "Boletim", "Protocolo", "Situação", "Área", "Bairro", "Problema", "Observação", "Aberto Por"]
        nova_linha = [protocolo_id, data_abertura, boletim, protocolo, situacao, area, bairro, problema, observacao, aberto_por]

        borda = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        centralizado = Alignment(horizontal="center")
        fonte_negrito = Font(bold=True)
        fundo_azul_claro = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")

        indices_centralizar = [1, 2, 3, 4, 5, 6, 10]  # Índices 1-based: Data, Boletim, Protocolo, Situação, Área, Aberto Por

        # Verifica se o arquivo existe
        if not caminho_excel.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Protocolos"
            ws.append(cabeçalhos)

            # Estiliza o cabeçalho
            for col in range(1, len(cabeçalhos) + 1):
                celula = ws.cell(row=1, column=col)
                celula.border = borda
                celula.font = fonte_negrito
                celula.fill = fundo_azul_claro
                if col in indices_centralizar:
                    celula.alignment = centralizado

            wb.save(caminho_excel)

        # Carrega o workbook existente e planilha
        wb = load_workbook(caminho_excel)
        ws = wb["Protocolos"]

        # Adiciona a nova linha
        ws.append(nova_linha)
        nova_linha_idx = ws.max_row

        # Aplica bordas e centraliza se necessário
        for col_num in range(1, len(nova_linha) + 1):
            cell = ws.cell(row=nova_linha_idx, column=col_num)
            cell.border = borda
            if col_num in indices_centralizar:
                cell.alignment = centralizado

        # Salva
        wb.save(caminho_excel)

        messagebox.showinfo("Sucesso", "Protocolo adicionado com sucesso!")

        # Atualiza a Treeview
        self.exibir_protocolos()

        # Seleciona o protocolo recém-adicionado na Treeview
        for item in self.treeview_protocolos.get_children():
            valores = self.treeview_protocolos.item(item, 'values')
            if int(valores[0]) == protocolo_id:  # coluna "ID" está na posição 0
                self.treeview_protocolos.selection_set(item)  # seleciona a linha
                self.treeview_protocolos.focus(item)  # foca a linha
                self.treeview_protocolos.see(item)  # rola até a linha
                break

        # Limpar os campos após a adição
        self.entry_boletim.delete(0, tk.END)
        self.entry_protocolo.delete(0, tk.END)
        self.combo_situacao.set("Situação (selecione)")
        self.combo_area.set("Área (selecione)")
        self.combo_problema.set("Problema (selecione)")
        #Corrige o bug do placeholder txt do campo obs desaparecer
        self.entry_observacoes.delete(0, tk.END)
        self.entry_observacoes.configure(placeholder_text="Observações(se houver)")

    def exibir_protocolos(self):
        caminho_banco = Path("sis_crm/database/db/sistema_protocolos.db")
        conn = sqlite3.connect(caminho_banco)  # Alterado o nome do banco
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM protocolos")
        protocolos = cursor.fetchall()
        conn.close()

        # Limpar a Treeview antes de exibir os dados novamente
        self.treeview_protocolos.delete(*self.treeview_protocolos.get_children())

        for protocolo in protocolos:
            self.treeview_protocolos.insert("", "end", values=protocolo)

    #MECANISMOS DE CONSULTA SQL NA TABELA
    # ATUALIZA A TABELA CONFORME O QUE É DIGITADO NA BARRA DE PESQUISA GERAL
    def atualizar_pesquisa_geral(self, event=None):
        pesquisa = self.entry_pesquisa_principal.get().strip()
        caminho_banco = Path("sis_crm/database/db/sistema_protocolos.db")
        conn = sqlite3.connect(caminho_banco)  # Alterado o nome do banco
        cursor = conn.cursor()

        if pesquisa.lower() == "centro":
            # Quando pesquisa for exatamente "Centro", busca apenas no bairro com igualdade
            cursor.execute('''
            SELECT id, data_abertura, boletim, protocolo, situacao, area, bairro, problema, observacao, aberto_por
            FROM protocolos
            WHERE bairro = ?
        ''', (pesquisa,))
        else:
        # Pesquisa genérica com LIKE para os outros campos
            cursor.execute('''
            SELECT id, data_abertura, boletim, protocolo, situacao, area, bairro, problema, observacao, aberto_por
            FROM protocolos
            WHERE boletim LIKE ?
            OR data_abertura LIKE ?
            OR situacao LIKE ?
            OR area LIKE ?
            OR problema LIKE ?
            OR observacao LIKE ?
            OR bairro LIKE ?
            OR aberto_por LIKE ?
            ''', (
            f'%{pesquisa}%', f'%{pesquisa}%', f'%{pesquisa}%', 
            f'%{pesquisa}%', f'%{pesquisa}%', f'%{pesquisa}%', 
            f'%{pesquisa}%', f'%{pesquisa}%'
            ))

        rows = cursor.fetchall()
        self.treeview_protocolos.delete(*self.treeview_protocolos.get_children())

        for row in rows:
            self.treeview_protocolos.insert("", "end", values=row)

        conn.close()

    # ATUALIZA A TABELA CONFORME O QUE É DIGITADO NA BARRA DE PESQUISA POR PROTOCOLO
    def atualizar_pesquisa_protocolo(self, event=None):
        pesquisa_protocolo = self.entry_pesquisa_id.get()
        caminho_banco = Path("sis_crm/database/db/sistema_protocolos.db")
        conn = sqlite3.connect(caminho_banco)  # Alterado o nome do banco
        cursor = conn.cursor()

        cursor.execute('''
        SELECT id, data_abertura, boletim, protocolo, situacao, area, bairro, problema, observacao, aberto_por
        FROM protocolos
        WHERE protocolos.id LIKE ? OR protocolos.protocolo LIKE ?
        ''', (f'%{pesquisa_protocolo}%', f'%{pesquisa_protocolo}%'))

        rows = cursor.fetchall()
        self.treeview_protocolos.delete(*self.treeview_protocolos.get_children())

        for row in rows:
            self.treeview_protocolos.insert("", "end", values=row)

        conn.close()

    #EXCLUIR PROTOCOLO DA TABELA
    def excluir_protocolo(self, event=None):
        item_selecionado = self.treeview_protocolos.focus()

        if not item_selecionado:
            messagebox.showwarning("Atenção", "Selecione um protocolo para excluir.")
            return

        valores = self.treeview_protocolos.item(item_selecionado, 'values')
        protocolo_id = valores[0]  # Pega o ID da primeira coluna

        confirmar = messagebox.askyesno("Confirmação", "Deseja excluir este protocolo? Esta ação é irreversível.", icon="warning")

        if confirmar:
            try:
                # Remove da Treeview imediatamente (feedback visual rápido)
                self.treeview_protocolos.delete(item_selecionado)
                
                # Executa a exclusão no banco de dados em uma thread separada
                threading.Thread(
                    target=self._excluir_protocolo_background,
                    args=(protocolo_id,),
                    daemon=True
                ).start()
                
                messagebox.showinfo("Sucesso", "Protocolo excluido com êxito.")

            except Exception as e:
                messagebox.showerror("Erro", f"Ocorreu um erro ao excluir: {e}")

    def _excluir_protocolo_background(self, protocolo_id):
        try:
            # Exclusão no banco de dados
            caminho_banco = Path("sis_crm/database/db/sistema_protocolos.db")
            conn = sqlite3.connect(caminho_banco)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM protocolos WHERE id = ?", (protocolo_id,))
            conn.commit()
            conn.close()

            # Exclusão no Excel (em outra thread para não bloquear)
            threading.Thread(
                target=self._remover_do_excel,
                args=(protocolo_id,),
                daemon=True
            ).start()

        except Exception as e:
            # Usar invoke para atualizar a UI na thread principal
            self.after(0, lambda: messagebox.showerror("Erro", f"Falha ao excluir protocolo: {e}"))

    def _remover_do_excel(self, protocolo_id):
        try:
            caminho_excel = Path("sis_crm/database/backup/backup.xlsx")
            if not caminho_excel.exists():
                return

            wb = load_workbook(caminho_excel)
            ws = wb["Protocolos"]

            # Encontrar e remover a linha com o ID correspondente
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                if str(row[0].value) == str(protocolo_id):
                    ws.delete_rows(idx)
                    break

            wb.save(caminho_excel)

        except Exception as e:
            self.after(0, lambda: print(f"Erro ao remover do Excel: {e}"))
    
    #MECANISMO DE EDIÇÃO DE DADOS
    def editar_protocolo(self, event=None):
        self.edicao_ativa = False  # Indica se o botão "Editar" foi pressionado

        # Bind do double click
        self.treeview_protocolos.bind("<Double-1>", self.editar_protocolo)

    def iniciar_edicao_se_modo_ativo(self, event):
        if self.edicao_ativa:
            self.edicao_ativa = False  # Desativa após primeiro clique
            self.editar_protocolo(event)

    def mostrar_calendario_ajustado(self, widget):
        # Aguarda o calendário abrir antes de reposicionar
        def ajustar_posicao():
            try:
                top = widget._top_cal  # TopLevel do calendário
                if top is not None:
                    # Posiciona acima do campo
                    root_x = widget.winfo_rootx()
                    root_y = widget.winfo_rooty()
                    top.geometry(f"+{root_x}+{root_y - 200}")
            except Exception as e:
                print(f"[Erro ao posicionar calendário]: {e}")

        # Abre o calendário de forma segura
        widget.after(150, lambda: (widget.drop_down(), ajustar_posicao()))

    #Validar o protocolo ex. DEN-000000-2025
    def validar_protocolo_edicao(self, event):
        """Valida o campo de protocolo durante a edição"""
        widget = event.widget
        valor = widget.get()
    
        # Remove espaços e converte para maiúsculas
        novo_valor = valor.replace(" ", "").upper()
    
        # Verifica se o valor foi alterado (havia espaços)
        if novo_valor != valor:
            widget.delete(0, tk.END)
            widget.insert(0, novo_valor)
            return
    
        # Verifica se excedeu o tamanho máximo (16 caracteres)
        if len(novo_valor) > 16:
            widget.delete(16, tk.END)
            return
    
        # Verifica caracteres inválidos (permite letras, números e hífen)
        for char in novo_valor:
            if not (char.isalnum() or char == '-'):
                widget.delete(len(novo_valor)-1, tk.END)
                return

    def editar_protocolo(self, event=None):
        """Edita um protocolo existente na tabela"""
        # Identifica o item e coluna clicada
        if event is not None:
            item_id = self.treeview_protocolos.identify_row(event.y)
            coluna = self.treeview_protocolos.identify_column(event.x)
        else:
            item_id = self.treeview_protocolos.focus()
            coluna = "#2"  # padrão para edição via botão

        if not item_id:
            messagebox.showwarning("Atenção", "Selecione um protocolo para editar.")
            return

        if coluna == '#1' or coluna == '#10':  # ID e Aberto por não podem ser editados
            return

        col_index = int(coluna.replace('#', '')) - 1
        col_nome = self.treeview_protocolos["columns"][col_index]
        x, y, largura, altura = self.treeview_protocolos.bbox(item_id, coluna)

        valor_atual = self.treeview_protocolos.set(item_id, col_nome)

        def salvar_edicao(event=None):
            """Salva as alterações feitas no protocolo"""
            try:
                if col_nome == "Data Abertura":
                    novo_valor = widget.get_date().strftime('%d/%m/%Y')
                else:
                    novo_valor = widget.get() if hasattr(widget, 'get') else widget.cget('text')
                
                self.treeview_protocolos.set(item_id, col_nome, novo_valor)
            
                # Fecha o widget de edição se ainda existir
                if widget.winfo_exists():
                    widget.destroy()

                # Atualiza o banco de dados
                valores = self.treeview_protocolos.item(item_id, 'values')
                id_protocolo = valores[0]

                colunas_db = {
                "Data Abertura": "data_abertura",
                "Boletim": "boletim",
                "Protocolo": "protocolo",
                "Situação": "situacao",
                "Área": "area",
                "Bairro": "bairro",
                "Problema": "problema",
                "Observação": "observacao"
                }

                if col_nome in colunas_db:
                    coluna_db = colunas_db[col_nome]

                    caminho_banco = Path("sis_crm/database/db/sistema_protocolos.db")
                    conn = sqlite3.connect(caminho_banco)
                    cursor = conn.cursor()
                    cursor.execute(f"UPDATE protocolos SET {coluna_db} = ? WHERE id = ?", (novo_valor, id_protocolo))
                    conn.commit()
                    conn.close()

                    threading.Thread(
                    target=self.atualizar_backup_excel,
                    args=(id_protocolo, col_nome, novo_valor),
                    daemon=True
                    ).start()

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar alteração: {e}")

        # Configuração do widget de edição
        if col_nome == "Data Abertura":
            # Calendário em PT-BR
            widget = DateEntry(
            self.treeview_protocolos,
            date_pattern='dd/mm/yyyy',
            locale='pt_BR',
            state='readonly',
            )
        
            try:
                dia, mes, ano = map(int, valor_atual.split('/'))
                widget.set_date(datetime(ano, mes, dia))
            except ValueError:
                pass
            
            self.mostrar_calendario_ajustado(widget)
            
            # Configura eventos
            def handle_calendar_return(event):
                if widget._top_cal.winfo_exists():
                    salvar_edicao()
        
            widget._calendar.bind("<Return>", handle_calendar_return)
            widget.bind("<Return>", salvar_edicao)  # <- aqui
            widget.bind("<<DateEntrySelected>>", lambda e: salvar_edicao())
            widget.bind("<<DateEntrySelected>>", lambda e: salvar_edicao())

        elif col_nome == "Protocolo":
            # Campo de protocolo com validação
            widget = ttk.Entry(self.treeview_protocolos)
            widget.insert(0, valor_atual)
            widget.bind("<KeyRelease>", self.validar_protocolo_edicao)
            widget.bind("<Return>", salvar_edicao)
            widget.bind("<FocusOut>", salvar_edicao)

        elif col_nome in ["Situação", "Área", "Problema"]:
            # Combobox readonly para campos de seleção
            opcoes = {
            "Situação": ["Aberto", "Reaberto", "Em andamento", "Em retorno", "Resolvido"],
            "Área": ["Sudoeste", "Centro-Sul", "Centro-Norte", "Noroeste", "Norte", "Leste"],
            "Problema": [
                "Focos de dengue", "Imóveis abandonados", "Acúmulo de lixo e materiais inservíveis",
                "Mato alto/entulho", "Animais Peçonhentos", "Pombos", "Morcegos",
                "Infestação de Roedores", "Carrapatos, pulgas e outros ectoparasitas",
                "Caramujos", "Criação de animais rurais em zona urbana",
                "Outras Reclamações de Saúde Pública"
            ]
            }
            widget = ttk.Combobox(
            self.treeview_protocolos,
            values=opcoes[col_nome],
            state="readonly"
            )
            widget.set(valor_atual)
            widget.bind("<<ComboboxSelected>>", salvar_edicao)
            widget.bind("<Return>", salvar_edicao)

        else:
            # Campo de texto normal para outros campos
            widget = ttk.Entry(self.treeview_protocolos)
            widget.insert(0, valor_atual)
            widget.bind("<Return>", salvar_edicao)
            widget.bind("<FocusOut>", salvar_edicao)

        # Posiciona o widget de edição
        widget.place(x=x, y=y, width=largura, height=altura)
        widget.focus()


    def atualizar_backup_excel(self, id_protocolo, col_nome, novo_valor):
        try:
            caminho_excel = Path("sis_crm/database/backup/backup.xlsx")
            if not caminho_excel.exists():
                return

            wb = load_workbook(caminho_excel)
            ws = wb["Protocolos"]

            for row in ws.iter_rows(min_row=2, values_only=False):
                if str(row[0].value) == str(id_protocolo):
                    idx_excel = list(self.treeview_protocolos["columns"]).index(col_nome)
                    row[idx_excel].value = novo_valor
                    
                    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
                    row[idx_excel].border = borda
                    if (idx_excel + 1) in [1, 2, 3, 4, 5, 6, 10]:
                        row[idx_excel].alignment = Alignment(horizontal="center")

            wb.save(caminho_excel)
        except Exception as e:
            print(f"Erro ao atualizar Excel: {e}")

    #função do botão que reinicia totalmente o sistema
    def recarregar_aplicacao(self):
        """Recarrega todos os componentes da aplicação sem fechar a janela principal"""
        # Destruir todos os frames existentes
        for widget in self.root.winfo_children():
            widget.destroy()
    
        # Reconstruir toda a interface
        self.__init__(self.root)

    def reiniciar_sistema(self):
        """Recarrega a aplicação sem fechar a janela principal"""
        self.recarregar_aplicacao()

    def confirmar_reinicio(self):
        resposta = messagebox.askyesno("Reiniciar", "Deseja realmente recarregar o sistema?")
        if resposta:
            self.reiniciar_sistema()

    #JANELA DE INFORMAÇÕES
    def abrir_janela_info(self):
        janela_info = ctk.CTkToplevel(self.root)
        janela_info.title("Informações do Sistema")
    
        # Definir tamanho e comportamento da janela
        janela_info.geometry("500x550")
        janela_info.resizable(False, False)
        janela_info.grab_set()
    
        # Centralizar a janela
        self.centralizar_janela(janela_info)
    
        # Configurar layout
        janela_info.grid_columnconfigure(0, weight=1)
        janela_info.grid_rowconfigure(1, weight=1)
    
        # Frame principal
        frame_principal = ctk.CTkFrame(janela_info, corner_radius=10)
        frame_principal.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        frame_principal.grid_columnconfigure(0, weight=1)
    
        # Título
        label_titulo = ctk.CTkLabel(
        frame_principal, 
        text="Informações do Sistema",
        font=("Poppins", 18, "bold"),
        anchor="center"
        )
        label_titulo.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
    
        # Informações estáticas
        self.icone_dev = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/dev.png"), size=(40, 40))
        ctk.CTkLabel(frame_principal, 
        text=" Desenvolvido por: Gabriel Henrique Constantino\n Estagiário - DZ/SES/PMS",
        image=self.icone_dev,
        compound="left",
        font=("Poppins", 14),
        anchor="w",
        justify="left"  # Isso alinha o texto à esquerda para múltiplas linhas
        ).grid(row=1, column=0, padx=20, pady=(10, 0), sticky="w")

        # Linha divisória (altura 2px)
        linha_divisoria = ctk.CTkFrame(
        frame_principal,
        height=2,
        fg_color="white",  # Cor branca
        border_width=0
        )
        linha_divisoria.grid(row=2, column=0, padx=20, pady=(10, 10), sticky="ew")
        
        self.icone_email = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/mail.png"), size=(24, 24))
        ctk.CTkLabel(frame_principal, 
        text=" Contato: gabrielconstantinogh@outlook.com",
        image=self.icone_email,
        font=("Poppins", 14), 
        compound="left",
        anchor="w",
        justify="left"
        ).grid(row=3, column=0, padx=20, sticky="w")
    
        # Links clicáveis 

        # Linha do LinkedIn
        icone_linkedin = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/linkedin.png"), size=(25, 25))
        ctk.CTkLabel(frame_principal,
        text=" LinkedIn:",
        image=icone_linkedin,
        compound="left",
        font=("Poppins", 14),
        anchor="w",
        justify="left"
        ).grid(row=4, column=0, padx=20, sticky="w")

        def abrir_linkedin():
            webbrowser.open_new("www.linkedin.com/in/gabrielhconstantino2")

        link_linkedin = ctk.CTkLabel(
        frame_principal, 
        text="linkedin.com/in/gabrielhconstantino2", 
        font=("Poppins", 14), 
        anchor="w",
        cursor="hand2",
        text_color="#1E90FF"
        )
        link_linkedin.grid(row=4, column=0, padx=110, sticky="w")  # Ajuste o padx para alinhamento
        link_linkedin.bind("<Button-1>", lambda e: abrir_linkedin())

        # Linha do GitHub
        self.icone_github = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/github.png"), size=(28, 28))
        ctk.CTkLabel(frame_principal, 
        text="Repositório:",
        image=self.icone_github, 
        compound="left",
        font=("Poppins", 14),
        anchor="w",
        justify="left"
        ).grid(row=5, column=0, padx=20, sticky="w")

        def abrir_github():
            webbrowser.open_new("https://github.com/gabrielhenriqueconstantino/sis_registro_crm_dz_sorocaba")

        link_github = ctk.CTkLabel(
        frame_principal, 
        text="github.com/sis_registro_crm_dz_sorocaba", 
        font=("Poppins", 14), 
        anchor="w",
        cursor="hand2",
        text_color="#1E90FF"
        )
        link_github.grid(row=5, column=0, padx=130, sticky="w")
        link_github.bind("<Button-1>", lambda e: abrir_github())

        # Linha da Documentação
        self.icone_documentacao = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/readme.png"), size=(25, 25))
        ctk.CTkLabel(frame_principal, 
        text=" Documentação:", 
        image=self.icone_documentacao,
        compound="left",
        font=("Arial", 14),
        anchor="w",
        justify="left"
        ).grid(row=6, column=0, padx=20, sticky="w")

        def abrir_docs():
            webbrowser.open_new("https://github.com/gabrielhenriqueconstantino/sis_registro_crm_dz_sorocaba/blob/main/README.md")

        link_docs = ctk.CTkLabel(
        frame_principal, 
        text="github.com/sis_registro_crm_dz_sorocaba/README.md", 
        font=("Poppins", 14), 
        anchor="w",
        cursor="hand2",
        text_color="#1E90FF"
        )
        link_docs.grid(row=6, column=0, padx=150, sticky="w")
        link_docs.bind("<Button-1>", lambda e: abrir_docs())
    
        # Informações estáticas restantes
        self.icone_versao = ctk.CTkImage(Image.open("sis_crm/img/icons/icons_app/version.png"), size=(25, 25))
        ctk.CTkLabel(frame_principal, 
        text=" Versão: 1.0.0 | 05/05/2025",
        image=self.icone_versao, 
        compound="left",
        font=("Poppins", 14),
        anchor="w",
        justify="left"
        ).grid(row=7, column=0, padx=20, pady=(0, 10), sticky="w")
    
        ctk.CTkLabel(frame_principal, 
        text="©2025 Sorocaba - \nEste software é livre e pode ser utilizado\n e modificado por setores da Prefeitura de Sorocaba.", 
        font=("Poppins", 14), 
        anchor="center"
        ).grid(row=8, column=0, padx=20, pady=(10, 20), sticky="ew")
    
        # Frame dos botões
        frame_botoes = ctk.CTkFrame(janela_info, fg_color="transparent")
        frame_botoes.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="ew")
        frame_botoes.grid_columnconfigure(0, weight=1)

        # Botão de fechar - TAMANHO AUMENTADO
        botao_fechar = ctk.CTkButton(
        frame_botoes, 
        text="Fechar", 
        command=janela_info.destroy,
        fg_color="#2b2b2b",
        hover_color="#3a3a3a",
        width=250,  
        height=40,  
        font=("Poppins", 14, "bold"),
        border_width=1,
        border_color="white"  # Fonte um pouco maior
        )
        botao_fechar.grid(row=0, column=0, pady=10)
    
        # Efeito hover para os links
        def on_enter(e, label):
            label.configure(font=("Poppins", 14, "underline"))
    
        def on_leave(e, label):
            label.configure(font=("Poppins", 14))
    
        for link in [link_linkedin, link_github, link_docs]:
            link.bind("<Enter>", lambda e, lbl=link: on_enter(e, lbl))
            link.bind("<Leave>", lambda e, lbl=link: on_leave(e, lbl))
            
        #Animação suave na abertura e fechamente da janela de informações   
        def animar_alpha(janela, inicio, fim, passos=10, intervalo=30, on_complete=None):
            delta = (fim - inicio) / passos
            atual = inicio

            def passo():
                nonlocal atual
                atual += delta
                if (delta > 0 and atual >= fim) or (delta < 0 and atual <= fim):
                    atual = fim
                    janela.attributes("-alpha", atual)
                    if on_complete:
                        on_complete()
                    return
                janela.attributes("-alpha", atual)
                janela.after(intervalo, passo)

            janela.attributes("-alpha", atual)
            passo()

        def fechar_janela_com_fade():
            animar_alpha(janela_info, 1.0, 0.0, on_complete=janela_info.destroy)

        # Substituir o botão "Fechar"
        botao_fechar.configure(command=fechar_janela_com_fade)

        # Inicia com fade in
        animar_alpha(janela_info, 0.0, 1.0)


    def centralizar_janela(self, janela):
        """Centraliza uma janela na tela"""
        janela.update_idletasks()
        width = janela.winfo_width()
        height = janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (width // 2)
        y = (janela.winfo_screenheight() // 2) - (height // 2)
        janela.geometry(f'+{x}+{y}')

    #JANELA DE ANÁLISE DE DADOS (está no arquivo analise_dados.py)
    def abrir_janela_analise(self):
        if self.janela_analise is None or not self.janela_analise.winfo_exists():
            self.janela_analise = JanelaAnalise(master=self.root)

            # Deixa a própria JanelaAnalise lidar com foco e efeito suave

            # Define ação ao fechar
            self.janela_analise.protocol("WM_DELETE_WINDOW", self.fechar_janela_analise)
        else:
            # Se já existe, traz para frente e foca
            self.janela_analise.lift()
            self.janela_analise.focus_force()
    
    def fechar_janela_analise(self):
        if self.janela_analise is not None:
            self.janela_analise.destroy()
            self.janela_analise = None


#INICIA O PROGRAMA
if __name__ == "__main__":
    root = ctk.CTk()
    app = SistemaProtocolos(root)
    root.mainloop()