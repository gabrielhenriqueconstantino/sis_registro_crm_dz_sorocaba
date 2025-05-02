"""
===========================================================================================
SCRIPT DE BACKUP DE DADOS DO SISTEMA DE PROTOCOLOS PARA EXCEL COM FORMATAÇÃO PERSONALIZADA
===========================================================================================

Este script tem como objetivo principal realizar um *backup completo e formatado* dos dados
armazenados no banco de dados SQLite do sistema de protocolos, salvando-os em um arquivo
Excel (.xlsx) com formatação visual e organizacional adequada para leitura, impressão ou
compartilhamento com outros setores da administração pública.

---------------------
FUNÇÃO DESTE SCRIPT:
---------------------
1. Detecta automaticamente a resolução da tela, útil para adaptar interfaces futuras.
2. Conecta-se ao banco de dados principal do sistema (`sistema_protocolos.db`) e extrai
   todos os registros existentes da tabela `protocolos`.
3. Cria uma nova planilha Excel chamada `backup.xlsx` dentro do diretório `database/backup/`.
4. Define e aplica estilos visuais à planilha, como:
   - Cabeçalhos com fundo azul claro e texto em negrito.
   - Bordas finas ao redor de todas as células.
   - Alinhamento centralizado para campos relevantes como datas, boletins, protocolos, etc.
5. Insere os dados do banco de dados na planilha e aplica a formatação mencionada.
6. Salva o arquivo Excel no local definido e imprime no terminal uma mensagem de sucesso.

------------------------
POR QUE O BACKUP EXCEL?
------------------------
O backup em Excel serve como uma *cópia externa e legível* dos dados do sistema, que pode ser:
- Utilizada em auditorias ou revisões externas.
- Compartilhada facilmente com gestores, equipes administrativas ou outros departamentos.
- Consultada de forma rápida mesmo sem acesso ao sistema principal.
- Impressa ou convertida em PDF, mantendo a estrutura visual organizada.
- Armazenada como um histórico de segurança adicional fora do banco de dados original.

Além disso, o Excel é um formato amplamente aceito e familiar, o que facilita a leitura por
usuários que não têm conhecimento técnico sobre bancos de dados.

-----------------------
ESTRUTURA DA PLANILHA:
-----------------------
A planilha gerada contém as seguintes colunas:
- ID
- Data Abertura
- Boletim
- Protocolo
- Situação
- Área
- Bairro
- Problema
- Observação
- Aberto Por

Cada linha representa um protocolo registrado no sistema, e os dados são automaticamente
extraídos diretamente da tabela do banco de dados, garantindo fidelidade à base oficial.

Este script pode ser agendado, executado manualmente ou incorporado como parte do processo
de exportação e backup regular do sistema de protocolos.

===========================================================================================
"""


from screeninfo import get_monitors

def obter_resolucao():
    monitor = get_monitors()[0]  # Pega o primeiro monitor
    return monitor.width, monitor.height

largura, altura = obter_resolucao()
print(f'Sua resolução é {largura}x{altura}')

import sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Caminhos
caminho_banco = Path("database/db/sistema_protocolos.db")
caminho_excel = Path("database/backup/backup.xlsx")
caminho_excel.parent.mkdir(parents=True, exist_ok=True)

# Conectar ao banco de dados
conn = sqlite3.connect(caminho_banco)
cursor = conn.cursor()

cursor.execute("SELECT * FROM protocolos")
dados = cursor.fetchall()
conn.close()

# Criar novo arquivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Protocolos"

# Estilos
fundo_azul_claro = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
fonte_negrito = Font(bold=True)
borda_fina = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
centralizado = Alignment(horizontal="center")

# Cabeçalhos
cabecalhos = ["ID", "Data Abertura", "Boletim", "Protocolo", "Situação", "Área", "Bairro", "Problema", "Observação", "Aberto Por"]
ws.append(cabecalhos)

# Índices dos campos a centralizar (baseado na lista acima)
indices_centralizar = [1, 2, 3, 4, 5, 6, 10]  # Python é 0-indexado

# Aplicar estilo ao cabeçalho
for col in range(1, len(cabecalhos) + 1):
    celula = ws.cell(row=1, column=col)
    celula.fill = fundo_azul_claro
    celula.font = fonte_negrito
    celula.border = borda_fina
    if col in indices_centralizar:
        celula.alignment = centralizado

# Inserir dados e aplicar bordas e alinhamento
for linha_idx, linha in enumerate(dados, start=2):
    for col_idx, valor in enumerate(linha, start=1):
        celula = ws.cell(row=linha_idx, column=col_idx, value=valor)
        celula.border = borda_fina
        if col_idx in indices_centralizar:
            celula.alignment = centralizado

# Salvar arquivo
wb.save(caminho_excel)
print("Backup com formatação e campos centralizados criado com sucesso!")